from __future__ import annotations

import filecmp
import json
import random
import re
import shutil
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

try:
    import requests
except ImportError:
    requests = None

try:
    import streamlit as st
except ImportError:
    st = None

try:
    import ai_config
except ImportError:
    ai_config = None


PROJECT_DIR = Path(__file__).resolve().parent
DATABASE_DIR = PROJECT_DIR / "Voc_Database"
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"
DEFAULT_BOOK_NAME = "默认词库"

WORD_COLUMNS = [
    "ID",
    "English",
    "Phonetic",
    "Chinese",
    "Part_Of_Speech",
    "TOEFL_Context",
    "Example_EN",
    "Example_CN",
    "Collocations",
    "Synonyms",
    "Antonyms",
    "Memory_Tip",
    "TOEFL_Writing_Use",
    "Mastery",
    "Next_Review",
    "Last_Review",
    "Correct_Count",
    "Wrong_Count",
    "Forget_Level",
    "Note",
    "Source",
    "AI_Status",
    "Created_At",
    "Updated_At",
]
INFO_FIELDS = [
    "Phonetic",
    "Chinese",
    "Part_Of_Speech",
    "TOEFL_Context",
    "Example_EN",
    "Example_CN",
    "Collocations",
    "Synonyms",
    "Antonyms",
    "Memory_Tip",
    "TOEFL_Writing_Use",
]
TEXT_COLUMNS = [
    "English",
    "Phonetic",
    "Chinese",
    "Part_Of_Speech",
    "TOEFL_Context",
    "Example_EN",
    "Example_CN",
    "Collocations",
    "Synonyms",
    "Antonyms",
    "Memory_Tip",
    "TOEFL_Writing_Use",
    "Next_Review",
    "Last_Review",
    "Forget_Level",
    "Note",
    "Source",
    "AI_Status",
    "Created_At",
    "Updated_At",
]
INT_COLUMNS = ["ID", "Mastery", "Correct_Count", "Wrong_Count"]
AI_STATUS_VALUES = {"pending", "done", "failed"}


class VocabDataError(Exception):
    pass


class DeepSeekError(Exception):
    pass


def config_value(name: str, default: str = "") -> str:
    if ai_config is None:
        return default
    return clean_text(getattr(ai_config, name, default))


def now_text() -> str:
    return datetime.now().strftime(DATE_FORMAT)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        cleaned = {str(key): clean_text(item) for key, item in value.items()}
        cleaned = {key: item for key, item in cleaned.items() if item}
        return json.dumps(cleaned, ensure_ascii=False) if cleaned else ""
    if isinstance(value, (list, tuple, set)):
        parts = []
        for item in value:
            item_text = clean_text(item)
            if item_text:
                parts.append(item_text)
        return "; ".join(parts)
    if not isinstance(value, (str, bytes)) and hasattr(value, "tolist"):
        try:
            return clean_text(value.tolist())
        except Exception:
            pass
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def safe_folder_name(name: str) -> str:
    name = clean_text(name) or DEFAULT_BOOK_NAME
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:80] or DEFAULT_BOOK_NAME


def normalize_ids(series: pd.Series) -> list[int]:
    raw_ids = pd.to_numeric(series, errors="coerce")
    used: set[int] = set()
    next_id = 1
    ids: list[int] = []
    for raw_id in raw_ids:
        candidate = int(raw_id) if pd.notna(raw_id) and int(raw_id) > 0 else None
        if candidate is not None and candidate not in used:
            ids.append(candidate)
            used.add(candidate)
            next_id = max(next_id, candidate + 1)
            continue
        while next_id in used:
            next_id += 1
        ids.append(next_id)
        used.add(next_id)
        next_id += 1
    return ids


def normalize_datetime_column(df: pd.DataFrame, column: str) -> list[str]:
    values = pd.to_datetime(df[column], errors="coerce")
    return [value.strftime(DATE_FORMAT) if pd.notna(value) else "" for value in values]


def review_heat_profile(wrong_count: int) -> dict[str, str]:
    if wrong_count >= 3:
        return {
            "label": "高频遗忘：3 次及以上",
            "card_background": "rgba(185, 28, 28, 0.14)",
            "card_border": "rgba(185, 28, 28, 0.42)",
            "badge_bg": "rgba(185, 28, 28, 0.16)",
            "badge_text": "#991b1b",
            "table_bg": "#fee2e2",
            "excel_fill": "FEE2E2",
            "excel_font": "991B1B",
        }
    if wrong_count == 2:
        return {
            "label": "重复遗忘：2 次",
            "card_background": "rgba(234, 88, 12, 0.12)",
            "card_border": "rgba(234, 88, 12, 0.34)",
            "badge_bg": "rgba(234, 88, 12, 0.14)",
            "badge_text": "#c2410c",
            "table_bg": "#ffedd5",
            "excel_fill": "FFEDD5",
            "excel_font": "C2410C",
        }
    if wrong_count == 1:
        return {
            "label": "第一次忘记：1 次",
            "card_background": "rgba(217, 119, 6, 0.10)",
            "card_border": "rgba(217, 119, 6, 0.28)",
            "badge_bg": "rgba(217, 119, 6, 0.12)",
            "badge_text": "#b45309",
            "table_bg": "#fef3c7",
            "excel_fill": "FEF3C7",
            "excel_font": "B45309",
        }
    return {
        "label": "目前还没有忘记记录",
        "card_background": "#ffffff",
        "card_border": "rgba(49, 51, 63, .16)",
        "badge_bg": "rgba(71, 85, 105, 0.08)",
        "badge_text": "#475569",
        "table_bg": "#f8fafc",
        "excel_fill": "F8FAFC",
        "excel_font": "475569",
    }


def has_completed_info(row: pd.Series | dict[str, object]) -> bool:
    chinese = clean_text(row.get("Chinese", ""))
    return bool(chinese and chinese != "待补充")


def normalize_ai_status(row: pd.Series | dict[str, object]) -> str:
    status = clean_text(row.get("AI_Status", "")).lower()
    if status in AI_STATUS_VALUES:
        return status
    return "done" if has_completed_info(row) else "pending"


def normalize_words(df: pd.DataFrame) -> pd.DataFrame:
    for column in WORD_COLUMNS:
        if column not in df.columns:
            df[column] = 0 if column in INT_COLUMNS else ""
    if df.empty:
        return df[WORD_COLUMNS]

    for column in TEXT_COLUMNS:
        df[column] = df[column].map(clean_text)
    for column in INT_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0).astype(int)
        df[column] = df[column].clip(lower=0)

    df["ID"] = normalize_ids(df["ID"])
    df["Forget_Level"] = df["Wrong_Count"].map(lambda count: review_heat_profile(int(count))["label"])
    df["Next_Review"] = normalize_datetime_column(df, "Next_Review")
    df["Last_Review"] = normalize_datetime_column(df, "Last_Review")
    timestamp = now_text()
    df["Created_At"] = df["Created_At"].replace("", timestamp)
    df["Updated_At"] = df["Updated_At"].replace("", timestamp)
    df["AI_Status"] = df.apply(normalize_ai_status, axis=1)
    extra_columns = [column for column in df.columns if column not in WORD_COLUMNS]
    return df[WORD_COLUMNS + extra_columns]


def word_info(row: pd.Series | dict[str, object]) -> dict[str, str]:
    return {field: clean_text(row.get(field, "")) for field in INFO_FIELDS}


def style_word_table(df: pd.DataFrame) -> pd.io.formats.style.Styler:
    def row_style(row: pd.Series) -> list[str]:
        wrong_count = int(row.get("Wrong_Count", 0))
        profile = review_heat_profile(wrong_count)
        styles: list[str] = []
        for column in row.index:
            if column in {"English", "Wrong_Count", "Forget_Level"}:
                styles.append(
                    f"background-color: {profile['table_bg']}; color: {profile['badge_text']}; font-weight: 600;"
                )
            else:
                styles.append("")
        return styles

    return df.style.apply(row_style, axis=1)


def apply_excel_review_heat(path: Path, df: pd.DataFrame) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    target_columns = [headers.get(name) for name in ("English", "Wrong_Count", "Forget_Level")]
    target_columns = [column for column in target_columns if column is not None]

    for row_number, (_, row) in enumerate(df.iterrows(), start=2):
        wrong_count = int(row.get("Wrong_Count", 0))
        profile = review_heat_profile(wrong_count)
        fill = PatternFill(fill_type="solid", fgColor=profile["excel_fill"])
        font = Font(color=profile["excel_font"], bold=wrong_count >= 3)
        for column_number in target_columns:
            cell = worksheet.cell(row=row_number, column=column_number)
            cell.fill = fill
            cell.font = font

    if "Forget_Level" in headers:
        worksheet.column_dimensions[worksheet.cell(row=1, column=headers["Forget_Level"]).column_letter].width = 18

    workbook.save(path)


def word_payload(row: pd.Series) -> dict[str, object]:
    payload = word_info(row)
    payload.update(
        {
            "ID": int(row.get("ID", 0)),
            "English": clean_text(row.get("English", "")),
            "Mastery": int(row.get("Mastery", 0)),
            "Next_Review": clean_text(row.get("Next_Review", "")),
            "Last_Review": clean_text(row.get("Last_Review", "")),
            "Correct_Count": int(row.get("Correct_Count", 0)),
            "Wrong_Count": int(row.get("Wrong_Count", 0)),
            "Forget_Level": clean_text(row.get("Forget_Level", "")),
            "Note": clean_text(row.get("Note", "")),
            "Source": clean_text(row.get("Source", "")),
            "AI_Status": normalize_ai_status(row),
        }
    )
    return payload


class AIClient:
    def __init__(self) -> None:
        self.provider = config_value("AI_PROVIDER", "deepseek").lower()
        self.mode = "openai_compatible"

        if self.provider == "openai":
            self.api_key = config_value("OPENAI_API_KEY")
            self.base_url = config_value("OPENAI_BASE_URL", "https://api.openai.com/v1").rstrip("/")
            self.model = config_value("OPENAI_MODEL", "gpt-4.1-mini")
            self.label = "OpenAI"
        elif self.provider == "gemini":
            self.api_key = config_value("GEMINI_API_KEY")
            self.base_url = config_value("GEMINI_BASE_URL", "https://generativelanguage.googleapis.com/v1beta").rstrip("/")
            self.model = config_value("GEMINI_MODEL", "gemini-2.0-flash")
            self.mode = "gemini"
            self.label = "Gemini"
        elif self.provider == "custom_openai_compatible":
            self.api_key = config_value("CUSTOM_OPENAI_COMPATIBLE_API_KEY")
            self.base_url = config_value("CUSTOM_OPENAI_COMPATIBLE_BASE_URL").rstrip("/")
            self.model = config_value("CUSTOM_OPENAI_COMPATIBLE_MODEL")
            self.label = "OpenAI-compatible"
        else:
            self.provider = "deepseek"
            self.api_key = config_value("DEEPSEEK_API_KEY")
            self.base_url = config_value("DEEPSEEK_BASE_URL", "https://api.deepseek.com").rstrip("/")
            self.model = config_value("DEEPSEEK_MODEL", "deepseek-chat")
            self.label = "DeepSeek"

    def available(self) -> bool:
        return bool(self.api_key and self.model and self.base_url) and requests is not None

    def status_text(self) -> str:
        if self.available():
            return f"{self.label} API 已配置"
        return f"{self.label} API 尚未完整配置"

    def enrich_word(self, word: str) -> dict[str, str]:
        prompt = f"""
你要为一名正在备考 TOEFL iBT 的中文母语学习者生成英语单词学习卡片。
目标单词："{word}"

只返回 JSON，不要 Markdown，不要解释。JSON 字段必须包含：
phonetic, chinese, part_of_speech, toefl_context, example_en, example_cn,
collocations, synonyms, antonyms, memory_tip, toefl_writing_use

内容要求：
1. chinese：用简洁中文解释核心 TOEFL 常见义项，优先学术阅读、听力、口语、写作中常见含义。
2. part_of_speech：写常见词性，可以包含多个。
3. toefl_context：说明这个词在 TOEFL 阅读/听力/写作/口语里常见的语境或题材。
4. example_en：写一个自然、学术、适合 TOEFL 的英文例句，不要太口语化。
5. example_cn：准确翻译 example_en。
6. collocations：给 3-5 个 TOEFL 写作或学术语境常用搭配。
7. synonyms / antonyms：给常见替换词和反义词，适合写作替换。
8. memory_tip：用中文给一个简短记忆提示。
9. toefl_writing_use：说明这个词在 TOEFL 写作/口语中如何使用，给一个短语级模板。
10. 不确定或不适合的字段返回空字符串。
11. 所有字段的值都必须是字符串，不要返回数组或对象；多个项目用分号分隔。
""".strip()

        content = self._complete(
            messages=[
                {"role": "system", "content": "你是一个严谨的 TOEFL 英语词汇教练。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            json_mode=True,
        )

        try:
            data = json.loads(content)
        except Exception as exc:
            raise DeepSeekError(f"{self.label} 返回内容不是可解析的 JSON。") from exc

        return {
            "Phonetic": clean_text(data.get("phonetic", "")),
            "Chinese": clean_text(data.get("chinese", "")),
            "Part_Of_Speech": clean_text(data.get("part_of_speech", "")),
            "TOEFL_Context": clean_text(data.get("toefl_context", "")),
            "Example_EN": clean_text(data.get("example_en", "")),
            "Example_CN": clean_text(data.get("example_cn", "")),
            "Collocations": clean_text(data.get("collocations", "")),
            "Synonyms": clean_text(data.get("synonyms", "")),
            "Antonyms": clean_text(data.get("antonyms", "")),
            "Memory_Tip": clean_text(data.get("memory_tip", "")),
            "TOEFL_Writing_Use": clean_text(data.get("toefl_writing_use", "")),
        }

    def chat(self, messages: list[dict[str, str]], book_name: str) -> str:
        system_prompt = f"""
你是一个 TOEFL 备考学习助手，正在帮助中文母语学习者使用本地单词本复习。
当前单词本：{book_name}

回答要求：
1. 优先围绕 TOEFL 阅读、听力、口语、写作场景解释。
2. 如果用户问单词，请给出简洁中文解释、常见 TOEFL 语境、例句和可用于写作/口语的表达。
3. 如果用户问作文或口语，请给结构化建议，但不要写得冗长。
4. 回答可以中英结合，但解释应以中文为主。
5. 不要声称已经修改 Excel；你只能聊天，不能直接改数据库。
""".strip()

        payload_messages = [{"role": "system", "content": system_prompt}]
        payload_messages.extend(messages[-12:])
        return self._complete(payload_messages, temperature=0.35, json_mode=False)

    def _complete(
        self,
        messages: list[dict[str, str]],
        temperature: float,
        json_mode: bool,
    ) -> str:
        if not self.api_key:
            raise DeepSeekError(f"还没有填写 {self.label} API Key。")
        if requests is None:
            raise DeepSeekError("缺少 requests，请先运行：pip install -r requirements.txt")
        if self.mode == "gemini":
            return self._gemini_complete(messages, temperature, json_mode)
        return self._openai_compatible_complete(messages, temperature, json_mode)

    def _openai_compatible_complete(
        self,
        messages: list[dict[str, str]],
        temperature: float,
        json_mode: bool,
    ) -> str:
        body: dict[str, object] = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
        }
        if json_mode:
            body["response_format"] = {"type": "json_object"}
        response = requests.post(
            f"{self.base_url}/chat/completions",
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            },
            json=body,
            timeout=60,
        )
        if response.status_code >= 400:
            raise DeepSeekError(f"{self.label} 请求失败：HTTP {response.status_code}")
        try:
            return clean_text(response.json()["choices"][0]["message"]["content"])
        except Exception as exc:
            raise DeepSeekError(f"{self.label} 返回内容无法解析。") from exc

    def _gemini_complete(
        self,
        messages: list[dict[str, str]],
        temperature: float,
        json_mode: bool,
    ) -> str:
        system_parts = []
        contents = []
        for message in messages:
            role = message.get("role", "user")
            content = clean_text(message.get("content", ""))
            if not content:
                continue
            if role == "system":
                system_parts.append({"text": content})
            else:
                contents.append(
                    {
                        "role": "model" if role == "assistant" else "user",
                        "parts": [{"text": content}],
                    }
                )

        generation_config: dict[str, object] = {"temperature": temperature}
        if json_mode:
            generation_config["response_mime_type"] = "application/json"

        body: dict[str, object] = {
            "contents": contents,
            "generationConfig": generation_config,
        }
        if system_parts:
            body["systemInstruction"] = {"parts": system_parts}

        response = requests.post(
            f"{self.base_url}/models/{self.model}:generateContent?key={self.api_key}",
            headers={"Content-Type": "application/json"},
            json=body,
            timeout=60,
        )
        if response.status_code >= 400:
            raise DeepSeekError(f"{self.label} 请求失败：HTTP {response.status_code}")

        try:
            parts = response.json()["candidates"][0]["content"]["parts"]
            return clean_text("".join(part.get("text", "") for part in parts))
        except Exception as exc:
            raise DeepSeekError(f"{self.label} 返回内容无法解析。") from exc


class VocabStore:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.root.mkdir(parents=True, exist_ok=True)
        if not self.book_records():
            self.create_book(DEFAULT_BOOK_NAME)

    def book_records(self) -> list[dict[str, str]]:
        records: list[dict[str, str]] = []
        for path in sorted(self.root.iterdir()):
            if path.is_dir() and not path.name.startswith("."):
                meta = self._read_meta(path)
                records.append({"name": meta["name"], "folder": path.name, "path": str(path)})
        return records

    def book_names(self) -> list[str]:
        return [record["name"] for record in self.book_records()]

    def book_path(self, book_name: str) -> Path:
        for record in self.book_records():
            if record["name"] == book_name:
                return Path(record["path"])
        raise VocabDataError("词库不存在。")

    def words_path(self, book_name: str) -> Path:
        return self.book_path(book_name) / "words.xlsx"

    def backups_path(self, book_name: str) -> Path:
        return self.book_path(book_name) / "backups"

    def _read_meta(self, book_path: Path) -> dict[str, str]:
        meta_path = book_path / "book.json"
        meta = {}
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                meta = {}
        timestamp = now_text()
        return {
            "name": clean_text(meta.get("name", book_path.name)) or book_path.name,
            "created_at": clean_text(meta.get("created_at", timestamp)) or timestamp,
            "updated_at": clean_text(meta.get("updated_at", timestamp)) or timestamp,
        }

    def _write_meta(self, book_path: Path, name: str, created_at: str | None = None) -> None:
        existing = self._read_meta(book_path) if (book_path / "book.json").exists() else {}
        meta = {
            "name": name,
            "created_at": created_at or existing.get("created_at") or now_text(),
            "updated_at": now_text(),
        }
        (book_path / "book.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    def _unique_folder(self, name: str) -> Path:
        base = safe_folder_name(name)
        candidate = self.root / base
        index = 2
        while candidate.exists():
            candidate = self.root / f"{base}_{index}"
            index += 1
        return candidate

    def create_book(self, book_name: str) -> None:
        book_name = clean_text(book_name)
        if not book_name:
            raise VocabDataError("词库名不能为空。")
        if book_name in self.book_names():
            raise VocabDataError("这个词库名已经存在。")
        book_path = self._unique_folder(book_name)
        book_path.mkdir(parents=True, exist_ok=False)
        (book_path / "backups").mkdir(exist_ok=True)
        self._write_meta(book_path, book_name)
        self.save_words(book_name, pd.DataFrame(columns=WORD_COLUMNS))

    def rename_book(self, old_name: str, new_name: str) -> None:
        new_name = clean_text(new_name)
        if not new_name:
            raise VocabDataError("新词库名不能为空。")
        if new_name != old_name and new_name in self.book_names():
            raise VocabDataError("这个词库名已经存在。")
        old_path = self.book_path(old_name)
        target = self.root / safe_folder_name(new_name)
        if target.exists() and target != old_path:
            target = self._unique_folder(new_name)
        if target != old_path:
            old_path.rename(target)
        self._write_meta(target, new_name)

    def load_words(self, book_name: str) -> pd.DataFrame:
        path = self.words_path(book_name)
        if not path.exists() or path.stat().st_size == 0:
            df = pd.DataFrame(columns=WORD_COLUMNS)
            self.save_words(book_name, df)
            return df
        try:
            df = pd.read_excel(path, engine="openpyxl")
        except Exception as exc:
            raise VocabDataError(f"无法读取 {path}。请确认它是正常的 .xlsx 文件。") from exc
        df = normalize_words(df)
        self.save_words(book_name, df)
        return df

    def save_words(self, book_name: str, df: pd.DataFrame) -> None:
        path = self.words_path(book_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        df = normalize_words(df)
        try:
            df.to_excel(path, index=False, engine="openpyxl")
            apply_excel_review_heat(path, df)
        except PermissionError as exc:
            raise VocabDataError(f"保存失败：请先关闭 Excel 中打开的 {path.name}。") from exc
        except Exception as exc:
            raise VocabDataError(f"保存 {path.name} 时出错：{exc}") from exc

    def duplicate_in_book(self, book_name: str, english: str) -> bool:
        df = self.load_words(book_name)
        normalized = english.strip().lower()
        return df["English"].astype(str).str.strip().str.lower().eq(normalized).any()

    def cached_info(self, english: str) -> dict[str, str] | None:
        normalized = english.strip().lower()
        for book in self.book_names():
            df = self.load_words(book)
            matches = df[df["English"].astype(str).str.strip().str.lower() == normalized]
            for _, row in matches.iterrows():
                if clean_text(row.get("Chinese", "")) and clean_text(row.get("Chinese", "")) != "待补充":
                    return word_info(row)
        return None

    def add_word(
        self,
        book_name: str,
        english: str,
        info: dict[str, str] | None,
        note: str,
        source: str,
        ai_status: str | None = None,
    ) -> dict[str, object]:
        english = english.strip()
        if not english:
            raise VocabDataError("英文单词不能为空。")
        if self.duplicate_in_book(book_name, english):
            raise VocabDataError("当前词库已经有这个单词了。")
        df = self.load_words(book_name)
        max_id = pd.to_numeric(df["ID"], errors="coerce").max() if not df.empty else pd.NA
        next_id = int(max_id) + 1 if pd.notna(max_id) else 1
        info = info or {}
        timestamp = now_text()
        status = clean_text(ai_status).lower()
        if status not in AI_STATUS_VALUES:
            status = "done" if has_completed_info(info) else "pending"
        row = {
            "ID": next_id,
            "English": english,
            "Phonetic": clean_text(info.get("Phonetic", "")),
            "Chinese": clean_text(info.get("Chinese", "")) or "待补充",
            "Part_Of_Speech": clean_text(info.get("Part_Of_Speech", "")),
            "TOEFL_Context": clean_text(info.get("TOEFL_Context", "")),
            "Example_EN": clean_text(info.get("Example_EN", "")),
            "Example_CN": clean_text(info.get("Example_CN", "")),
            "Collocations": clean_text(info.get("Collocations", "")),
            "Synonyms": clean_text(info.get("Synonyms", "")),
            "Antonyms": clean_text(info.get("Antonyms", "")),
            "Memory_Tip": clean_text(info.get("Memory_Tip", "")),
            "TOEFL_Writing_Use": clean_text(info.get("TOEFL_Writing_Use", "")),
            "Mastery": 0,
            "Next_Review": "",
            "Last_Review": "",
            "Correct_Count": 0,
            "Wrong_Count": 0,
            "Note": note.strip(),
            "Source": source,
            "AI_Status": status,
            "Created_At": timestamp,
            "Updated_At": timestamp,
        }
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        self.save_words(book_name, df)
        return row

    def ai_pending_ids(self, book_name: str, include_failed: bool = True) -> list[int]:
        df = self.load_words(book_name)
        valid = self.valid_indices(book_name)
        if not valid:
            return []
        statuses = df.loc[valid, "AI_Status"].map(clean_text).str.lower()
        wanted = {"pending", "failed"} if include_failed else {"pending"}
        mask = statuses.isin(wanted) | df.loc[valid, "Chinese"].map(clean_text).eq("待补充")
        return [int(df.at[index, "ID"]) for index in statuses[mask].index]

    def ai_pending_count(self, book_name: str, include_failed: bool = True) -> int:
        return len(self.ai_pending_ids(book_name, include_failed=include_failed))

    def word_by_id(self, book_name: str, word_id: int) -> pd.Series:
        df = self.load_words(book_name)
        matches = df.index[df["ID"] == int(word_id)].tolist()
        if not matches:
            raise VocabDataError("这个单词不存在，请重新读取 Excel。")
        return df.loc[matches[0]]

    def update_word_info(
        self,
        book_name: str,
        word_id: int,
        info: dict[str, str],
        source: str,
        ai_status: str = "done",
    ) -> dict[str, object]:
        df = self.load_words(book_name)
        matches = df.index[df["ID"] == int(word_id)].tolist()
        if not matches:
            raise VocabDataError("这个单词不存在，请重新读取 Excel。")
        index = matches[0]
        for field in INFO_FIELDS:
            value = clean_text(info.get(field, ""))
            if value:
                df.at[index, field] = value
        if not clean_text(df.at[index, "Chinese"]):
            df.at[index, "Chinese"] = "待补充"
        status = clean_text(ai_status).lower()
        df.at[index, "Source"] = source
        df.at[index, "AI_Status"] = status if status in AI_STATUS_VALUES else normalize_ai_status(df.loc[index])
        df.at[index, "Updated_At"] = now_text()
        self.save_words(book_name, df)
        return word_payload(df.loc[index])

    def mark_ai_status(self, book_name: str, word_id: int, ai_status: str, source: str | None = None) -> None:
        df = self.load_words(book_name)
        matches = df.index[df["ID"] == int(word_id)].tolist()
        if not matches:
            raise VocabDataError("这个单词不存在，请重新读取 Excel。")
        index = matches[0]
        status = clean_text(ai_status).lower()
        df.at[index, "AI_Status"] = status if status in AI_STATUS_VALUES else "pending"
        if source:
            df.at[index, "Source"] = source
        df.at[index, "Updated_At"] = now_text()
        self.save_words(book_name, df)

    def update_word_basic(self, book_name: str, word_id: int, english: str, note: str) -> dict[str, object]:
        english = clean_text(english)
        if not english:
            raise VocabDataError("英文单词不能为空。")

        df = self.load_words(book_name)
        matches = df.index[df["ID"] == int(word_id)].tolist()
        if not matches:
            raise VocabDataError("这个单词不存在，请重新读取 Excel。")
        index = matches[0]

        normalized = english.lower()
        duplicate_mask = df["English"].astype(str).str.strip().str.lower().eq(normalized)
        duplicate_mask.loc[index] = False
        if duplicate_mask.any():
            raise VocabDataError("当前词库已经有这个单词了。")

        old_english = clean_text(df.at[index, "English"])
        df.at[index, "English"] = english
        df.at[index, "Note"] = clean_text(note)
        if old_english.lower() != normalized:
            df.at[index, "Source"] = "manual_edit"
            df.at[index, "AI_Status"] = "pending"
        df.at[index, "Updated_At"] = now_text()
        self.save_words(book_name, df)
        return word_payload(df.loc[index])

    def valid_indices(self, book_name: str) -> list[int]:
        df = self.load_words(book_name)
        return df[df["English"].astype(str).str.strip().ne("")].index.tolist()

    def due_indices(self, book_name: str) -> list[int]:
        df = self.load_words(book_name)
        valid = self.valid_indices(book_name)
        if not valid:
            return []
        current = datetime.now()
        next_review = pd.to_datetime(df.loc[valid, "Next_Review"], errors="coerce")
        due_mask = next_review.isna() | (next_review <= current)
        return list(next_review[due_mask].index)

    def stats(self, book_name: str) -> dict[str, int]:
        df = self.load_words(book_name)
        valid = self.valid_indices(book_name)
        if not valid:
            return {"total": 0, "mastered": 0, "due": 0}
        mastery = pd.to_numeric(df.loc[valid, "Mastery"], errors="coerce").fillna(0)
        return {"total": len(valid), "mastered": int((mastery >= 3).sum()), "due": len(self.due_indices(book_name))}

    def next_due_word(self, book_name: str, exclude_id: int | None = None) -> dict[str, object] | None:
        df = self.load_words(book_name)
        due = self.due_indices(book_name)
        if exclude_id is not None and len(due) > 1:
            due = [index for index in due if int(df.at[index, "ID"]) != int(exclude_id)]
        if not due:
            return None
        return word_payload(df.loc[random.choice(due)])

    def answer(self, book_name: str, word_id: int, remembered: bool) -> None:
        df = self.load_words(book_name)
        matches = df.index[df["ID"] == int(word_id)].tolist()
        if not matches:
            raise VocabDataError("这个单词不存在，请重新读取 Excel。")
        index = matches[0]
        current = datetime.now()
        mastery = int(df.at[index, "Mastery"])
        correct_count = int(df.at[index, "Correct_Count"])
        wrong_count = int(df.at[index, "Wrong_Count"])
        if remembered:
            mastery += 1
            correct_count += 1
            if mastery == 1:
                next_review = current + timedelta(hours=12)
            elif mastery == 2:
                next_review = current + timedelta(days=1)
            else:
                next_review = current + timedelta(days=3)
        else:
            mastery = 0
            wrong_count += 1
            next_review = current + timedelta(minutes=5)
        df.at[index, "Mastery"] = mastery
        df.at[index, "Correct_Count"] = correct_count
        df.at[index, "Wrong_Count"] = wrong_count
        df.at[index, "Last_Review"] = current.strftime(DATE_FORMAT)
        df.at[index, "Next_Review"] = next_review.strftime(DATE_FORMAT)
        df.at[index, "Updated_At"] = current.strftime(DATE_FORMAT)
        self.save_words(book_name, df)

    def backup_if_changed(self, book_name: str) -> dict[str, object]:
        excel_path = self.words_path(book_name)
        if not excel_path.exists():
            return {"created": False, "reason": "没有可备份的 Excel。"}
        backup_dir = self.backups_path(book_name)
        backup_dir.mkdir(exist_ok=True)
        backups = sorted(backup_dir.glob("words_*.xlsx"))
        if backups and filecmp.cmp(excel_path, backups[-1], shallow=False):
            return {"created": False, "reason": "Excel 与最近备份一致，已跳过备份。"}
        target = backup_dir / f"words_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(excel_path, target)
        return {"created": True, "path": str(target), "reason": "已创建新备份。"}

    def book_details(self, book_name: str) -> dict[str, object]:
        book_path = self.book_path(book_name)
        excel_path = self.words_path(book_name)
        backup_dir = self.backups_path(book_name)
        backups = sorted(backup_dir.glob("words_*.xlsx")) if backup_dir.exists() else []
        df = self.load_words(book_name)
        modified = datetime.fromtimestamp(excel_path.stat().st_mtime).strftime(DATE_FORMAT) if excel_path.exists() else ""
        return {
            "book_path": str(book_path),
            "excel_path": str(excel_path),
            "backup_count": len(backups),
            "latest_backup": str(backups[-1]) if backups else "",
            "modified_at": modified,
            "dataframe": df,
        }

class App:
    def __init__(self) -> None:
        self.store = VocabStore(DATABASE_DIR)
        self.ai = AIClient()

    def add_word_fast(self, book_name: str, english: str, note: str) -> dict[str, object]:
        if self.store.duplicate_in_book(book_name, english):
            raise VocabDataError("当前词库已经有这个单词了。")
        info = self.store.cached_info(english)
        if info is not None:
            row = self.store.add_word(book_name, english, info, note, "local_cache", "done")
            return {"word": row, "info": word_info(row), "source": "local_cache", "warning": ""}
        row = self.store.add_word(book_name, english, None, note, "manual_pending", "pending")
        return {
            "word": row,
            "info": word_info(row),
            "source": "manual_pending",
            "warning": "已快速保存，AI 资料可稍后批量补全。",
        }

    def enrich_and_save(self, book_name: str, english: str, note: str) -> dict[str, object]:
        if self.store.duplicate_in_book(book_name, english):
            raise VocabDataError("当前词库已经有这个单词了。")
        info = self.store.cached_info(english)
        source = "local_cache"
        status = "done"
        warning = ""
        if info is None:
            source = self.ai.provider
            try:
                info = self.ai.enrich_word(english)
            except DeepSeekError as exc:
                info = None
                source = f"pending_{self.ai.provider}"
                status = "failed"
                warning = str(exc)
        row = self.store.add_word(book_name, english, info, note, source, status)
        return {"word": row, "info": word_info(row), "source": source, "warning": warning}

    def enrich_existing_word(self, book_name: str, word_id: int) -> dict[str, object]:
        row = self.store.word_by_id(book_name, word_id)
        english = clean_text(row.get("English", ""))
        if not english:
            raise VocabDataError("这个单词没有英文内容。")

        info = self.store.cached_info(english)
        source = "local_cache"
        if info is None:
            source = self.ai.provider
            try:
                info = self.ai.enrich_word(english)
            except DeepSeekError as exc:
                self.store.mark_ai_status(book_name, word_id, "failed", f"failed_{self.ai.provider}")
                raise exc
        updated = self.store.update_word_info(book_name, word_id, info, source, "done")
        return {"word": updated, "info": word_info(updated), "source": source, "warning": ""}

    def enrich_pending_words(self, book_name: str, limit: int) -> dict[str, object]:
        pending_ids = self.store.ai_pending_ids(book_name)
        selected_ids = pending_ids[: max(0, int(limit))]
        result: dict[str, object] = {"done": 0, "failed": 0, "skipped": 0, "messages": []}
        messages: list[str] = []
        for word_id in selected_ids:
            try:
                enriched = self.enrich_existing_word(book_name, word_id)
                word = enriched["word"]
                result["done"] = int(result["done"]) + 1
                messages.append(f"{word.get('English', '')}: done via {enriched['source']}")
            except (VocabDataError, DeepSeekError) as exc:
                result["failed"] = int(result["failed"]) + 1
                messages.append(f"ID {word_id}: failed - {exc}")
        result["skipped"] = max(0, len(pending_ids) - len(selected_ids))
        result["messages"] = messages
        return result


APP = App()


def running_in_streamlit() -> bool:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx(suppress_warning=True) is not None
    except Exception:
        return False


def init_state() -> None:
    if "page_radio" not in st.session_state:
        st.session_state.page_radio = "输入新词"
    if st.session_state.get("_go_review"):
        st.session_state.page_radio = "复习"
        del st.session_state["_go_review"]
    books = APP.store.book_names()
    if "current_book" not in st.session_state or st.session_state.current_book not in books:
        st.session_state.current_book = books[0]
    if "current_word" not in st.session_state:
        st.session_state.current_word = None
    if "answer_visible" not in st.session_state:
        st.session_state.answer_visible = False
    if "last_preview" not in st.session_state:
        st.session_state.last_preview = "最近一次补全结果会显示在这里。"
    if "last_added_id" not in st.session_state:
        st.session_state.last_added_id = None
    if "detail_message" not in st.session_state:
        st.session_state.detail_message = ""
    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = []


def apply_style() -> None:
    st.markdown(
        """
        <style>
        .block-container { max-width: 1100px; padding-top: 2rem; }
        .word-card {
            border: 1px solid rgba(49, 51, 63, .16);
            border-radius: 12px;
            padding: 28px;
            margin-top: 12px;
        }
        .word-face {
            font-size: 58px;
            line-height: 1.05;
            font-weight: 750;
            text-align: center;
            padding: 28px 8px;
            overflow-wrap: anywhere;
        }
        .small-note { color: #667085; font-size: 0.92rem; line-height: 1.55; }
        div[data-testid="stMetric"] {
            border: 1px solid rgba(49, 51, 63, .14);
            border-radius: 10px;
            padding: 10px 12px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def current_book() -> str:
    return st.session_state.current_book


def set_current_book(book_name: str) -> None:
    if book_name != st.session_state.current_book:
        st.session_state.current_book = book_name
        st.session_state.current_word = None
        st.session_state.answer_visible = False


def render_sidebar() -> None:
    with st.sidebar:
        st.header("单词本")
        books = APP.store.book_names()
        selected = st.selectbox("当前单词本", books, index=books.index(current_book()))
        set_current_book(selected)

        name = st.text_input("新建 / 改名")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("新建", use_container_width=True):
                try:
                    APP.store.create_book(name)
                    st.session_state.current_book = name.strip()
                    st.session_state.current_word = None
                    st.rerun()
                except VocabDataError as exc:
                    st.error(str(exc))
        with col2:
            if st.button("改名", use_container_width=True):
                try:
                    APP.store.rename_book(current_book(), name)
                    st.session_state.current_book = name.strip()
                    st.session_state.current_word = None
                    st.rerun()
                except VocabDataError as exc:
                    st.error(str(exc))

        st.divider()
        if APP.ai.available():
            st.success(APP.ai.status_text())
        else:
            st.warning(APP.ai.status_text())
        st.caption(f"数据库：{DATABASE_DIR}")


def render_metrics() -> None:
    stats = APP.store.stats(current_book())
    c1, c2, c3 = st.columns(3)
    c1.metric("总词数", stats["total"])
    c2.metric("已熟练", stats["mastered"])
    c3.metric("当前待复习", stats["due"])


def render_input_page() -> None:
    st.subheader("输入新词")
    pending_count = APP.store.ai_pending_count(current_book())
    st.caption(f"当前还有 {pending_count} 个单词等待 AI 补全。快速加入不会卡住背词节奏。")
    with st.form("add_word_form", clear_on_submit=True):
        english = st.text_input("英文单词")
        note = st.text_input("备注，可选")
        col_fast, col_full = st.columns(2)
        with col_fast:
            fast_submitted = st.form_submit_button("快速加入", use_container_width=True)
        with col_full:
            enrich_submitted = st.form_submit_button("加入并补全", use_container_width=True)

    if fast_submitted or enrich_submitted:
        if not english.strip():
            st.warning("请先输入英文单词。")
        else:
            try:
                if fast_submitted:
                    result = APP.add_word_fast(current_book(), english.strip(), note.strip())
                else:
                    with st.spinner(f"正在处理 {english.strip()}，会优先使用本地缓存。"):
                        result = APP.enrich_and_save(current_book(), english.strip(), note.strip())
                info = result["info"]
                lines = [f"单词：{english.strip()}", f"来源：{result['source']}"]
                if result["warning"]:
                    lines.append(f"提示：{result['warning']}")
                lines.extend([f"{field}：{info.get(field, '')}" for field in INFO_FIELDS])
                st.session_state.last_preview = "\n".join(lines)
                st.session_state.last_added_id = int(result["word"]["ID"])
                st.session_state.current_word = None
                st.session_state.answer_visible = False
                st.success(f"{english.strip()} 已保存到 words.xlsx。")
            except (VocabDataError, DeepSeekError) as exc:
                st.error(str(exc))

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("补全最近加入的单词", use_container_width=True):
            if st.session_state.last_added_id is None:
                st.warning("还没有最近加入的单词。")
            else:
                try:
                    with st.spinner("正在补全最近加入的单词..."):
                        result = APP.enrich_existing_word(current_book(), int(st.session_state.last_added_id))
                    info = result["info"]
                    lines = [f"单词：{result['word'].get('English', '')}", f"来源：{result['source']}"]
                    lines.extend([f"{field}：{info.get(field, '')}" for field in INFO_FIELDS])
                    st.session_state.last_preview = "\n".join(lines)
                    st.success("最近加入的单词已补全并保存。")
                except (VocabDataError, DeepSeekError) as exc:
                    st.error(str(exc))
    with c2:
        batch_limit = st.number_input("本次批量补全数量", min_value=1, max_value=50, value=10, step=1)

    if st.button("批量补全未完成单词", use_container_width=True):
        if pending_count == 0:
            st.info("当前没有等待补全的单词。")
        else:
            with st.spinner("正在批量补全，成功一个就会立刻写入 Excel..."):
                result = APP.enrich_pending_words(current_book(), int(batch_limit))
            messages = result["messages"]
            st.session_state.last_preview = "\n".join(messages) if messages else "没有需要补全的单词。"
            st.success(f"批量补全完成：成功 {result['done']} 个，失败 {result['failed']} 个，未处理 {result['skipped']} 个。")

    if st.button("停止输入，开始复习"):
        result = APP.store.backup_if_changed(current_book())
        st.toast(result["reason"])
        st.session_state._go_review = True
        st.rerun()

    st.text_area("最近一次补全结果", st.session_state.last_preview, height=260)


def answer_text(word: dict[str, object]) -> str:
    return "\n".join(
        [
            f"中文：{word.get('Chinese', '')}",
            f"音标：{word.get('Phonetic', '')}",
            f"词性：{word.get('Part_Of_Speech', '')}",
            f"TOEFL 语境：{word.get('TOEFL_Context', '')}",
            f"例句：{word.get('Example_EN', '')}",
            f"翻译：{word.get('Example_CN', '')}",
            f"搭配：{word.get('Collocations', '')}",
            f"近义：{word.get('Synonyms', '')}",
            f"反义：{word.get('Antonyms', '')}",
            f"记忆提示：{word.get('Memory_Tip', '')}",
            f"TOEFL 写作/口语使用：{word.get('TOEFL_Writing_Use', '')}",
            f"备注：{word.get('Note', '')}",
        ]
    )


def render_speech_button(word: str) -> None:
    import streamlit.components.v1 as components
    safe_word = json.dumps(word, ensure_ascii=False)
    components.html(
        f"""
        <button onclick='const u = new SpeechSynthesisUtterance({safe_word});
        u.lang = "en-US"; window.speechSynthesis.cancel(); window.speechSynthesis.speak(u);'
        style="border:0;border-radius:8px;background:#4b5563;color:white;min-height:40px;padding:0 16px;
        font:16px -apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;cursor:pointer;">发音</button>
        """,
        height=48,
    )


def ensure_word() -> None:
    if st.session_state.current_word is None:
        st.session_state.current_word = APP.store.next_due_word(current_book())
        st.session_state.answer_visible = False


def review_heat_style(wrong_count: int) -> dict[str, str]:
    profile = review_heat_profile(wrong_count)
    return {
        "background": profile["card_background"],
        "border": profile["card_border"],
        "badge_bg": profile["badge_bg"],
        "badge_text": profile["badge_text"],
        "label": profile["label"],
    }


def render_review_page() -> None:
    render_metrics()
    ensure_word()
    word = st.session_state.current_word

    if word is None:
        stats = APP.store.stats(current_book())
        if stats["total"] == 0:
            st.info("先输入几个单词吧。")
        else:
            st.success("今天的复习任务已完成。复习时间会按当前日期时间实时判断。")
        return

    wrong_count = int(word.get("Wrong_Count", 0))
    heat = review_heat_style(wrong_count)
    st.markdown(
        f'<div class="word-card" style="background:{heat["background"]};border-color:{heat["border"]};">',
        unsafe_allow_html=True,
    )
    st.markdown('<p class="small-note">先回忆中文、音标、例句和 TOEFL 使用场景，再显示答案。</p>', unsafe_allow_html=True)
    st.markdown(
        f'<div style="display:flex;justify-content:center;margin:8px 0 2px;">'
        f'<span style="display:inline-block;border-radius:999px;padding:6px 12px;'
        f'background:{heat["badge_bg"]};color:{heat["badge_text"]};font-size:14px;font-weight:600;">'
        f'{heat["label"]}</span></div>',
        unsafe_allow_html=True,
    )
    st.markdown(f'<div class="word-face">{word["English"]}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 1, 1])
    with c1:
        if st.button("显示答案", use_container_width=True):
            st.session_state.answer_visible = True
    with c2:
        render_speech_button(str(word["English"]))
    with c3:
        if st.button("换一个到期词", use_container_width=True):
            st.session_state.current_word = APP.store.next_due_word(current_book(), int(word["ID"]))
            st.session_state.answer_visible = False
            st.rerun()

    if st.session_state.answer_visible:
        st.text_area("答案", answer_text(word), height=320)
        f1, f2 = st.columns(2)
        with f1:
            if st.button("忘记", use_container_width=True):
                APP.store.answer(current_book(), int(word["ID"]), False)
                st.session_state.current_word = APP.store.next_due_word(current_book(), int(word["ID"]))
                st.session_state.answer_visible = False
                st.rerun()
        with f2:
            if st.button("认识", type="primary", use_container_width=True):
                APP.store.answer(current_book(), int(word["ID"]), True)
                st.session_state.current_word = APP.store.next_due_word(current_book(), int(word["ID"]))
                st.session_state.answer_visible = False
                st.rerun()


def render_detail_page() -> None:
    details = APP.store.book_details(current_book())
    render_metrics()
    st.subheader("单词本内部信息")
    if st.session_state.detail_message:
        st.success(st.session_state.detail_message)
        st.session_state.detail_message = ""
    st.write("单词本文件夹：", details["book_path"])
    st.write("Excel 文件：", details["excel_path"])
    st.write("最近修改：", details["modified_at"])
    st.write("备份数量：", details["backup_count"])
    st.write("最新备份：", details["latest_backup"] or "暂无")
    st.dataframe(style_word_table(details["dataframe"]), use_container_width=True, height=420)

    st.divider()
    st.subheader("修改单词")
    df = details["dataframe"]
    valid_df = df[df["English"].astype(str).str.strip().ne("")]
    if valid_df.empty:
        st.info("当前词库还没有可以修改的单词。")
        return

    st.caption("可以输入 ID 精确定位，也可以搜索拼错的英文片段。搜索到多个结果时，把目标 ID 填到左侧。")
    c1, c2 = st.columns([1, 2])
    with c1:
        raw_id = st.text_input("按 ID 定位", placeholder="例如：17")
    with c2:
        query = st.text_input("搜索英文单词", placeholder="输入拼错的单词或其中一部分")

    selected_id: int | None = None
    valid_ids = set(valid_df["ID"].astype(int).tolist())
    if raw_id.strip():
        try:
            candidate_id = int(raw_id.strip())
            if candidate_id in valid_ids:
                selected_id = candidate_id
            else:
                st.warning("这个 ID 不在当前词库里。")
        except ValueError:
            st.warning("ID 需要输入数字。")

    if query.strip():
        normalized_query = re.escape(query.strip())
        matches = valid_df[
            valid_df["English"].astype(str).str.contains(normalized_query, case=False, na=False, regex=True)
        ].sort_values("ID")
        if matches.empty:
            st.info("没有找到匹配的单词。")
        else:
            visible_columns = ["ID", "English", "Chinese", "Wrong_Count", "Forget_Level", "Note", "AI_Status"]
            st.dataframe(style_word_table(matches[visible_columns].head(30)), use_container_width=True)
            if selected_id is None:
                exact_matches = matches[matches["English"].astype(str).str.strip().str.lower() == query.strip().lower()]
                if len(exact_matches) == 1:
                    selected_id = int(exact_matches.iloc[0]["ID"])
                elif len(matches) == 1:
                    selected_id = int(matches.iloc[0]["ID"])
                else:
                    st.info("找到多个匹配结果，请把要修改的 ID 填到左侧。")

    if selected_id is None:
        st.info("先输入 ID，或搜索一个能唯一定位的单词。")
        return

    selected_row = APP.store.word_by_id(current_book(), selected_id)
    st.caption(f"正在修改：ID {selected_id} | {clean_text(selected_row.get('English', ''))}")

    with st.form("edit_word_form"):
        new_english = st.text_input("英文单词", value=clean_text(selected_row.get("English", "")))
        new_note = st.text_input("备注", value=clean_text(selected_row.get("Note", "")))
        submitted = st.form_submit_button("保存修改")

    if submitted:
        try:
            updated = APP.store.update_word_basic(current_book(), selected_id, new_english, new_note)
            st.session_state.current_word = None
            st.session_state.answer_visible = False
            st.session_state.detail_message = f"{updated['English']} 已保存到 words.xlsx。"
            st.rerun()
        except VocabDataError as exc:
            st.error(str(exc))


def render_extra_page() -> None:
    st.subheader("作文素材")
    st.info("这个地方在施工。。。")


def render_chat_page() -> None:
    st.subheader("AI 聊天")
    st.caption("这里可以问 TOEFL 单词、例句、写作表达和口语思路。聊天记录只保存在当前页面会话里，不会写入 Excel。")

    if not APP.ai.available():
        st.warning(f"{APP.ai.label} API 尚未完整配置。请先在 ai_config.py 里填写对应配置。")

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("清空聊天"):
            st.session_state.chat_messages = []
            st.rerun()
    with c2:
        st.caption(f"当前单词本：{current_book()}")

    for message in st.session_state.chat_messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    prompt = st.chat_input(f"问 {APP.ai.label}：比如 explain ubiquitous for TOEFL writing")
    if prompt:
        user_message = {"role": "user", "content": prompt}
        st.session_state.chat_messages.append(user_message)
        with st.chat_message("user"):
            st.markdown(prompt)

        with st.chat_message("assistant"):
            try:
                with st.spinner(f"{APP.ai.label} 正在思考..."):
                    answer = APP.ai.chat(st.session_state.chat_messages, current_book())
                st.markdown(answer)
                st.session_state.chat_messages.append({"role": "assistant", "content": answer})
            except DeepSeekError as exc:
                st.error(str(exc))


def render_app() -> None:
    st.set_page_config(page_title="TOEFL Excel 背单词", layout="wide")
    apply_style()
    init_state()
    render_sidebar()
    st.title("TOEFL Excel 背单词")
    page = st.radio(
        "页面",
        ["输入新词", "复习", "单词本详情", "AI 聊天", "扩展"],
        horizontal=True,
        key="page_radio",
        label_visibility="collapsed",
    )
    if page == "输入新词":
        render_input_page()
    elif page == "复习":
        render_review_page()
    elif page == "单词本详情":
        render_detail_page()
    elif page == "AI 聊天":
        render_chat_page()
    else:
        render_extra_page()


def main() -> None:
    if running_in_streamlit():
        if st is None:
            raise SystemExit("缺少 streamlit，请先运行：pip install -r requirements.txt")
        render_app()
        return
    print("这个版本使用 Streamlit 面板。请运行：")
    print("  ./run_app.sh")
    print(f"项目目录：{PROJECT_DIR}")
    print(f"数据库目录：{DATABASE_DIR}")


if __name__ == "__main__":
    main()
